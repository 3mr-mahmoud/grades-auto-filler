import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import cv2
import numpy as np
from utils.paper_extraction import extract_paper
from gradeSheet.extractCells import extract_cells
from gradeSheet.classifier.classifier import processCells
from gradeSheet.OCR import processCellsOCR
from gradeSheet.classifier.train import train
import joblib
import os
from bubbleSheet.final import process_bubble_sheet_gui

# Paths to save trained models
DIGITS_MODEL_PATH = 'digits_model.pkl'
SYMBOLS_MODEL_PATH = 'symbols_model.pkl'

# Function to train or load models
def train_or_load_models():
    if os.path.exists(DIGITS_MODEL_PATH) and os.path.exists(SYMBOLS_MODEL_PATH):
        # Load the models from disk if they exist
        digits_model = joblib.load(DIGITS_MODEL_PATH)
        symbols_model = joblib.load(SYMBOLS_MODEL_PATH)
    else:
        # Train the models if they don't exist
        digits_model = train('../../../Dataset/Training Set/digits_dataset2')
        symbols_model = train('../../../Dataset/Training Set/symbols_dataset')
        
        # Save the models to disk
        joblib.dump(digits_model, DIGITS_MODEL_PATH)
        joblib.dump(symbols_model, SYMBOLS_MODEL_PATH)

    return digits_model, symbols_model

# Load or train models
digits_model, symbols_model = train_or_load_models()

# Streamlit App Layout
st.title("Sheet Processor")

# User selection for sheet type
sheet_type = st.radio("Select the type of sheet to process:", ["Grade Sheet", "Bubble Sheet"])

if sheet_type == "Grade Sheet":
    # Upload image
    uploaded_file = st.file_uploader("Upload an Image of the Grade Sheet", type=["jpg", "jpeg", "png"])

    if uploaded_file:
        # Read the uploaded image
        file_bytes = uploaded_file.read()
        nparr = np.frombuffer(file_bytes, np.uint8)
        paper = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

        # Display the uploaded image
        st.image(paper, caption="Uploaded Grade Sheet", use_container_width=True)

        # Extract the paper (warped image)
        warped_image = extract_paper(paper)

        if warped_image is not None:
            st.image(warped_image, caption="Warped Paper", use_container_width=True)

            # Extract cells from the warped paper
            df, rows, cols = extract_cells(warped_image)

            # Select processing option
            option = st.radio("Select Processing Option", ["Use Model Classifier", "Use OCR"])

            # Process cells
            result = None
            if option == "Use Model Classifier":
                result = processCells(df, digits_model, symbols_model)
            else:
                result = processCellsOCR(df, symbols_model)

            # Drop unnecessary columns and clean up the result
            result = result.drop(result.columns[[1, 2]], axis=1)
            result = result.drop(index=0).reset_index(drop=True)
            result.columns = ["Student ID", "1", "2", "3"]

            st.write("Processed Results:")
            st.dataframe(result)

            # Highlight cells with "red" and export to Excel
            excel_path = "Grade sheet results.xlsx"
            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                result.to_excel(writer, index=False, sheet_name="Sheet1")
                workbook = writer.book
                worksheet = writer.sheets["Sheet1"]

                red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                rows, cols = result.shape

                for i in range(rows):
                    for j in range(cols):
                        value = result.iloc[i, j]
                        if isinstance(value, str) and value.strip().lower() == "red":
                            cell = worksheet.cell(row=i + 2, column=j + 1)
                            cell.value = ""
                            cell.fill = red_fill

            # Provide a download button for the Excel file
            st.success("Processing Complete! Download the results below:")
            with open(excel_path, "rb") as file:
                btn = st.download_button(
                    label="Download Excel",
                    data=file,
                    file_name="output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

elif sheet_type == "Bubble Sheet":
    uploaded_file = st.file_uploader("Upload an Image of the Bubble Sheet", type=["jpg", "jpeg", "png"])
    uploaded_answers = st.file_uploader("Upload the Answer Key", type=["txt"])

    if uploaded_file and uploaded_answers:
        st.write("Processing...")
        results_df = process_bubble_sheet_gui(uploaded_file, uploaded_answers)

        if isinstance(results_df, pd.DataFrame):
            st.write("Processing Complete!")
            st.dataframe(results_df)

            # Provide a download button for the results
            excel_path = "bubble_sheet_results.xlsx"
            results_df.to_excel(excel_path, index=False)

            with open(excel_path, "rb") as file:
                btn = st.download_button(
                    label="Download Results",
                    data=file,
                    file_name="bubble_sheet_results.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.error(f"Error: {results_df}")
